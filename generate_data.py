"""
GENERADOR BASE DE CONOCIMIENTO - DASHBOARD FARMACIAS

Lee dos archivos:
  - Dashboard.xlsx  → 5 hojas operativas, se actualiza cada hora via EasyMorph
  - Objetivos.xlsx  → archivo fijo, se reemplaza una vez por mes

Genera base_conocimiento.csv en la misma carpeta.

DEPENDENCIAS:
    pip install pandas openpyxl
"""

import pandas as pd
from openpyxl import load_workbook
import datetime
import calendar
import os

# ==============================================================================
# CONFIGURACION - AJUSTAR RUTAS
# ==============================================================================

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))

# ==============================================================================
# DESCARGA AUTOMATICA DESDE GOOGLE DRIVE
# ==============================================================================

DASHBOARD_FILE_ID = '1rTow4rq7UJL4Kuts-JdMLBS6AUqXcHXUrYOgEWj_fI4'
TOKEN_PATH        = os.path.join(BASE_DIR, "input", "token.json")
CREDS_PATH        = os.path.join(BASE_DIR, "input", "credentials.json")

def download_dashboard_from_drive(dest_path):
    """Descarga el Dashboard de Google Drive usando token OAuth guardado."""
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
        import io

        if not os.path.exists(TOKEN_PATH):
            print(f"  WARN: Sin token.json — usando archivo local existente.")
            print(f"        Corre setup_google_auth.py una vez para activar la descarga automatica.")
            return False

        creds = Credentials.from_authorized_user_file(TOKEN_PATH)

        # Refrescar token si venció
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(TOKEN_PATH, 'w') as f:
                f.write(creds.to_json())

        service  = build('drive', 'v3', credentials=creds)
        request  = service.files().export_media(
            fileId   = DASHBOARD_FILE_ID,
            mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

        with open(dest_path, 'wb') as f:
            f.write(buf.getvalue())

        print(f"  OK: Dashboard descargado desde Drive ({len(buf.getvalue())//1024} KB)")
        return True

    except Exception as e:
        print(f"  WARN: Error descargando de Drive: {e}")
        print(f"        Usando archivo local existente.")
        return False

PATH_DASH     = os.path.join(BASE_DIR, "input", "Dashboard (2).xlsx")
PATH_OBJ      = os.path.join(BASE_DIR, "input", "Objetivos Sucursales Abril 2026 - SIN UN KIOSCO.xlsx")
PATH_OUT      = os.path.join(BASE_DIR, "public", "data", "base_conocimiento.csv")
PATH_OUT_HORA = os.path.join(BASE_DIR, "public", "data", "horas_hoy.csv")
PATH_OUT_SEM  = os.path.join(BASE_DIR, "public", "data", "horas_semana_anterior.csv")

# ==============================================================================
# HELPERS
# ==============================================================================

def normalizar(v):
    return str(v).upper().strip()

# ==============================================================================
# 1. FACTURACION ACUMULADA DEL MES
# ==============================================================================

print("[0/8] Descargando Dashboard desde Google Drive...")
download_dashboard_from_drive(PATH_DASH)

print("[1/8] Leyendo facturacion acumulada del mes...")
df_mes = pd.read_excel(PATH_DASH, sheet_name='Facturacion Mes')
df_mes['Sucursal'] = df_mes['Nombre_Sucu'].apply(normalizar)
df_mes = df_mes.rename(columns={
    'sucursal':        'ID_Sucursal',
    'Tickets':         'Acum_Tickets',
    'Unidades':        'Acum_Unidades',
    'Neto':            'Acum_Neto',
    'Saldo_Cobertura': 'Acum_Cobertura',
    'Saldo_Cliente':   'Acum_Cliente',
})
df_mes = df_mes[[
    'ID_Sucursal', 'Sucursal',
    'Acum_Tickets', 'Acum_Unidades', 'Acum_Neto', 'Acum_Cobertura', 'Acum_Cliente',
]]

# ==============================================================================
# 2. FACTURACION DEL DIA + DETECCION AUTOMATICA DE FECHA
# ==============================================================================

print("[2/8] Leyendo facturacion del dia y detectando fecha...")
df_dia = pd.read_excel(PATH_DASH, sheet_name='Facturacion x Dia')
df_dia['Sucursal'] = df_dia['Nombre_Sucu'].apply(normalizar)

fecha_hoy      = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(df_dia['Fecha'].iloc[0]))
dia_actual     = fecha_hoy.day
dias_mes       = calendar.monthrange(fecha_hoy.year, fecha_hoy.month)[1]
dias_restantes = dias_mes - dia_actual
dias_restantes_safe = max(dias_restantes, 1)  # evita división por cero el último día del mes

print(f"    Fecha: {fecha_hoy} | Dia {dia_actual}/{dias_mes} | Dias restantes: {dias_restantes}")

df_dia = df_dia.rename(columns={
    'Tickets':         'Hoy_Tickets',
    'Unidades':        'Hoy_Unidades',
    'Neto':            'Hoy_Neto',
    'Saldo_Cobertura': 'Hoy_Cobertura',
    'Saldo_Cliente':   'Hoy_Cliente',
})
df_dia = df_dia[[
    'Sucursal', 'Hoy_Tickets', 'Hoy_Unidades', 'Hoy_Neto', 'Hoy_Cobertura', 'Hoy_Cliente',
]]

# ==============================================================================
# 3. SUCURSAL X HORA (HOY)
# ==============================================================================

print("[3/8] Leyendo datos por hora...")
df_hora = pd.read_excel(PATH_DASH, sheet_name='Sucursal x Hora')
df_hora['Sucursal'] = df_hora['Nombre_Sucu'].apply(normalizar)

ultima_franja_hoy = int(df_hora['Franja_Horaria'].max())
print(f"    Ultima franja disponible: {ultima_franja_hoy}hs")

hora_pico = (
    df_hora.loc[df_hora.groupby('Sucursal')['Neto'].idxmax()]
    [['Sucursal', 'Franja_Horaria', 'Neto']].copy()
)
hora_pico.columns = ['Sucursal', 'Hora_Pico_Hoy', 'Hora_Pico_Hoy_Neto']

ult_franja = df_hora.groupby('Sucursal')['Franja_Horaria'].max().reset_index()
ult_franja.columns = ['Sucursal', 'Ultima_Franja_Con_Datos']

neto_hora = df_hora.groupby('Sucursal').agg(
    Hora_Neto_Total=('Neto', 'sum'),
    Hora_Tickets_Total=('Tickets', 'sum'),
    Hora_Unidades_Total=('Unidades', 'sum'),
).reset_index()

df_hora_res = hora_pico.merge(ult_franja, on='Sucursal').merge(neto_hora, on='Sucursal')

# ==============================================================================
# 4. SEMANA ANTERIOR
#    Comparacion hasta la misma franja de ahora + total dia completo
# ==============================================================================

print("[4/8] Leyendo semana anterior...")
df_sem = pd.read_excel(PATH_DASH, sheet_name='Sucursal Semana Anterior')
df_sem['Sucursal'] = df_sem['Nombre_Sucu'].apply(normalizar)

sem_hasta = (
    df_sem[df_sem['Franja_Horaria'] <= ultima_franja_hoy]
    .groupby('Sucursal').agg(
        SemAnt_Neto_HastaAhora=('Neto', 'sum'),
        SemAnt_Tickets_HastaAhora=('Tickets', 'sum'),
        SemAnt_Unidades_HastaAhora=('Unidades', 'sum'),
    ).reset_index()
)
sem_total = df_sem.groupby('Sucursal').agg(
    SemAnt_Neto_DiaCompleto=('Neto', 'sum'),
    SemAnt_Tickets_DiaCompleto=('Tickets', 'sum'),
).reset_index()
df_sem_res = sem_hasta.merge(sem_total, on='Sucursal')

# ==============================================================================
# 5. ULTIMA HORA DE TICKET
# ==============================================================================

print("[5/8] Leyendo ultima hora de ticket...")
df_ult = pd.read_excel(PATH_DASH, sheet_name='Ultima Hora Ticket')
df_ult['Sucursal'] = df_ult['sucursal(2)'].apply(normalizar)
df_ult = df_ult.rename(columns={'ultima_hora': 'Ultima_Hora_Ticket'})[['Sucursal', 'Ultima_Hora_Ticket']]

# ==============================================================================
# 6. OBJETIVOS DESDE ARCHIVO SEPARADO
#    6a. Metas totales del mes (Entregable a Sucursales)
#    6b. Meta acumulada real dias 1 a hoy (Reporte Encargado)
# ==============================================================================

print("[6/8] Leyendo objetivos...")
wb_obj = load_workbook(PATH_OBJ, read_only=True)

# 6a. Metas totales del mes
ws_ent   = wb_obj['Entregable a Sucursales']
rows_ent = list(ws_ent.iter_rows(values_only=True))
data_ent = [r for r in rows_ent[6:] if r[1] and str(r[1]).lower() != 'total general']
df_obj   = pd.DataFrame(data_ent, columns=rows_ent[5])
df_obj   = df_obj[['FARMACIA', 'Meta 1 UN', 'Meta 1 $$', 'Meta 2 UN', 'Meta 2 $$', 'Meta 3 UN', 'Meta 3 $$']].copy()
df_obj.columns = [
    'Sucursal',
    'Obj_Meta1_UN_Mes',    'Obj_Meta1_Pesos_Mes',
    'Obj_Meta2_UN_Mes',    'Obj_Meta2_Pesos_Mes',
    'Obj_Meta3_UN_Mes',    'Obj_Meta3_Pesos_Mes',
]
df_obj['Sucursal'] = df_obj['Sucursal'].apply(normalizar)

# 6b. Meta acumulada real dias 1 a dia_actual
ws_enc   = wb_obj['REPORTE ENCARGADO']
rows_enc = list(ws_enc.iter_rows(values_only=True))
enc_data = []
for row in rows_enc[6:]:
    if row[2] is None or row[4] is None:
        continue
    if isinstance(row[2], datetime.datetime) and row[2].day <= dia_actual:
        enc_data.append({
            'Sucursal': normalizar(row[4]),
            'M1P': row[6]  or 0,
            'M2P': row[8]  or 0,
            'M3P': row[10] or 0,
            'M1U': row[5]  or 0,
        })
df_enc = pd.DataFrame(enc_data).groupby('Sucursal').agg(
    MetaAcum_Meta1_Pesos=('M1P', 'sum'),
    MetaAcum_Meta2_Pesos=('M2P', 'sum'),
    MetaAcum_Meta3_Pesos=('M3P', 'sum'),
    MetaAcum_Meta1_UN   =('M1U', 'sum'),
).reset_index()

# ==============================================================================
# 7. MERGE MAESTRO
# ==============================================================================

print("[7/8] Unificando y calculando metricas...")
df = (
    df_mes
    .merge(df_dia,      on='Sucursal', how='left')
    .merge(df_hora_res, on='Sucursal', how='left')
    .merge(df_sem_res,  on='Sucursal', how='left')
    .merge(df_ult,      on='Sucursal', how='left')
    .merge(df_obj,      on='Sucursal', how='left')
    .merge(df_enc,      on='Sucursal', how='left')
)

# ==============================================================================
# 8. COLUMNAS CALCULADAS
# ==============================================================================

# Contexto temporal
df['Ctx_Fecha_Hoy']            = str(fecha_hoy)
df['Ctx_Dia_Del_Mes']          = dia_actual
df['Ctx_Dias_Totales_Mes']     = dias_mes
df['Ctx_Dias_Restantes']       = dias_restantes
df['Ctx_Pct_Mes_Transcurrido'] = round(dia_actual / dias_mes * 100, 1)
df['Ctx_Ultima_Franja_Hora']   = ultima_franja_hoy

# Ritmo y proyeccion
df['Ritmo_Real_Diario']      = round(df['Acum_Neto'] / dia_actual, 0)
df['Proyeccion_Neto_FinMes'] = round(df['Acum_Neto'] + df['Ritmo_Real_Diario'] * dias_restantes, 0)
df['Proyeccion_UN_FinMes']   = round(df['Acum_Unidades'] + df['Acum_Unidades'] / dia_actual * dias_restantes, 0)

# Avance vs meta acumulada REAL (no division lineal del mes)
df['Avance_Pct_vs_MetaAcum_M1']    = round(df['Acum_Neto']     / df['MetaAcum_Meta1_Pesos'].replace(0, float('nan')) * 100, 1).fillna(0)
df['Avance_Pct_vs_MetaAcum_M2']    = round(df['Acum_Neto']     / df['MetaAcum_Meta2_Pesos'].replace(0, float('nan')) * 100, 1).fillna(0)
df['Avance_Pct_vs_MetaAcum_M3']    = round(df['Acum_Neto']     / df['MetaAcum_Meta3_Pesos'].replace(0, float('nan')) * 100, 1).fillna(0)
df['Avance_Pct_vs_MetaAcum_UN_M1'] = round(df['Acum_Unidades'] / df['MetaAcum_Meta1_UN'].replace(0, float('nan'))    * 100, 1).fillna(0)

# Superavit / deficit vs meta acumulada (+ bien, - mal)
df['Delta_vs_MetaAcum_M1'] = round(df['Acum_Neto'] - df['MetaAcum_Meta1_Pesos'], 0)
df['Delta_vs_MetaAcum_M2'] = round(df['Acum_Neto'] - df['MetaAcum_Meta2_Pesos'], 0)
df['Delta_vs_MetaAcum_M3'] = round(df['Acum_Neto'] - df['MetaAcum_Meta3_Pesos'], 0)

# Cuanto falta para cerrar cada meta del mes completo
df['Falta_Meta1_FinMes'] = round(df['Obj_Meta1_Pesos_Mes'] - df['Acum_Neto'], 0)
df['Falta_Meta2_FinMes'] = round(df['Obj_Meta2_Pesos_Mes'] - df['Acum_Neto'], 0)
df['Falta_Meta3_FinMes'] = round(df['Obj_Meta3_Pesos_Mes'] - df['Acum_Neto'], 0)

# Ritmo necesario en dias restantes para llegar a cada meta
df['Ritmo_Necesario_Meta1'] = round(df['Falta_Meta1_FinMes'].clip(lower=0) / dias_restantes_safe, 0)
df['Ritmo_Necesario_Meta2'] = round(df['Falta_Meta2_FinMes'].clip(lower=0) / dias_restantes_safe, 0)
df['Ritmo_Necesario_Meta3'] = round(df['Falta_Meta3_FinMes'].clip(lower=0) / dias_restantes_safe, 0)

# Variacion vs semana anterior (mismo tramo horario, comparacion justa)
df['Var_Pct_vs_SemAnt_HastaAhora'] = round(
    (df['Hoy_Neto'] - df['SemAnt_Neto_HastaAhora']) / df['SemAnt_Neto_HastaAhora'].replace(0, float('nan')) * 100, 1).fillna(0)

# Estado actual vs meta acumulada real
def estado_acumulado(row):
    if pd.isna(row.get('MetaAcum_Meta1_Pesos')): return 'SIN OBJETIVO'
    n = row['Acum_Neto']
    if   n >= row['MetaAcum_Meta3_Pesos']:        return 'ARRIBA META 3'
    elif n >= row['MetaAcum_Meta2_Pesos']:        return 'ARRIBA META 2'
    elif n >= row['MetaAcum_Meta1_Pesos']:        return 'ARRIBA META 1'
    elif row['Avance_Pct_vs_MetaAcum_M1'] >= 90: return 'CERCA META 1'
    else:                                         return 'DEBAJO META 1'
df['Estado_Acumulado'] = df.apply(estado_acumulado, axis=1)

# Meta proyectada a fin de mes
def meta_proyectada(row):
    p = row['Proyeccion_Neto_FinMes']
    if pd.isna(p) or pd.isna(row.get('Obj_Meta1_Pesos_Mes')): return 'SIN OBJETIVO'
    if   p >= row['Obj_Meta3_Pesos_Mes']: return 'META 3'
    elif p >= row['Obj_Meta2_Pesos_Mes']: return 'META 2'
    elif p >= row['Obj_Meta1_Pesos_Mes']: return 'META 1'
    else:                                 return 'SIN META'
df['Meta_Proyectada_FinMes'] = df.apply(meta_proyectada, axis=1)

# Semaforo
SEMAFORO = {
    'ARRIBA META 3': 'VERDE SOBRE META 3',
    'ARRIBA META 2': 'AMARILLO SOBRE META 2',
    'ARRIBA META 1': 'NARANJA SOBRE META 1',
    'CERCA META 1':  'NARANJA CERCA META 1',
    'DEBAJO META 1': 'ROJO DEBAJO META 1',
    'SIN OBJETIVO':  'SIN OBJETIVO',
}
df['Semaforo'] = df['Estado_Acumulado'].map(SEMAFORO)

# Alerta de inactividad
def alerta_inactividad(row):
    try:
        p = str(row['Ultima_Hora_Ticket']).split(':')
        h = int(p[0]) + int(p[1]) / 60
        return f'SIN TICKET DESDE {row["Ultima_Hora_Ticket"]}' if h < (ultima_franja_hoy - 1) else 'ACTIVA'
    except Exception:
        return 'SIN DATOS'
df['Alerta_Inactividad'] = df.apply(alerta_inactividad, axis=1)

# ==============================================================================
# ORDEN FINAL DE COLUMNAS (61 columnas)
# ==============================================================================

columnas_finales = [
    # Identificacion
    'ID_Sucursal', 'Sucursal',
    # Contexto temporal
    'Ctx_Fecha_Hoy', 'Ctx_Dia_Del_Mes', 'Ctx_Dias_Totales_Mes', 'Ctx_Dias_Restantes',
    'Ctx_Pct_Mes_Transcurrido', 'Ctx_Ultima_Franja_Hora',
    # Facturacion acumulada del mes
    'Acum_Tickets', 'Acum_Unidades', 'Acum_Neto', 'Acum_Cobertura', 'Acum_Cliente',
    # Facturacion de hoy
    'Hoy_Tickets', 'Hoy_Unidades', 'Hoy_Neto', 'Hoy_Cobertura', 'Hoy_Cliente',
    # Actividad por hora
    'Hora_Pico_Hoy', 'Hora_Pico_Hoy_Neto', 'Ultima_Franja_Con_Datos',
    'Hora_Neto_Total', 'Hora_Tickets_Total', 'Hora_Unidades_Total',
    'Ultima_Hora_Ticket', 'Alerta_Inactividad',
    # Semana anterior
    'SemAnt_Neto_HastaAhora', 'SemAnt_Tickets_HastaAhora', 'SemAnt_Unidades_HastaAhora',
    'SemAnt_Neto_DiaCompleto', 'SemAnt_Tickets_DiaCompleto', 'Var_Pct_vs_SemAnt_HastaAhora',
    # Objetivos totales del mes
    'Obj_Meta1_UN_Mes', 'Obj_Meta1_Pesos_Mes',
    'Obj_Meta2_UN_Mes', 'Obj_Meta2_Pesos_Mes',
    'Obj_Meta3_UN_Mes', 'Obj_Meta3_Pesos_Mes',
    # Meta acumulada real dias 1 a hoy
    'MetaAcum_Meta1_Pesos', 'MetaAcum_Meta2_Pesos', 'MetaAcum_Meta3_Pesos', 'MetaAcum_Meta1_UN',
    # Avance vs meta acumulada real
    'Avance_Pct_vs_MetaAcum_M1', 'Avance_Pct_vs_MetaAcum_M2',
    'Avance_Pct_vs_MetaAcum_M3', 'Avance_Pct_vs_MetaAcum_UN_M1',
    # Superavit / deficit (+ bien, - mal)
    'Delta_vs_MetaAcum_M1', 'Delta_vs_MetaAcum_M2', 'Delta_vs_MetaAcum_M3',
    # Cuanto falta para cerrar el mes
    'Falta_Meta1_FinMes', 'Falta_Meta2_FinMes', 'Falta_Meta3_FinMes',
    # Ritmo
    'Ritmo_Real_Diario', 'Ritmo_Necesario_Meta1', 'Ritmo_Necesario_Meta2', 'Ritmo_Necesario_Meta3',
    # Proyecciones
    'Proyeccion_Neto_FinMes', 'Proyeccion_UN_FinMes',
    # Conclusion
    'Estado_Acumulado', 'Meta_Proyectada_FinMes', 'Semaforo',
]
df = df[columnas_finales]

# ==============================================================================
# GUARDAR
# ==============================================================================

print("[8/8] Guardando archivos de salida...")

# Leer hojas de beneficios
df_clientes = pd.read_excel(PATH_DASH, sheet_name='Acumulado Clientes')
df_nominados = pd.read_excel(PATH_DASH, sheet_name='Ticket nominados')
alta_clientes     = int(df_clientes['TotalClientes'].iloc[0]) if not df_clientes.empty else 0
pct_nominados     = float(df_nominados['PctNominados'].iloc[0]) if not df_nominados.empty else 0
total_tickets_nom = int(df_nominados['TicketsNominados'].iloc[0]) if not df_nominados.empty else 0
total_tickets_nom_base = int(df_nominados['TotalTickets'].iloc[0]) if not df_nominados.empty else 0

# base_conocimiento.csv
df.fillna('').to_csv(PATH_OUT, index=False, encoding='utf-8-sig')

# horas_hoy.csv - detalle hora x sucursal del dia actual
df_hora_out = df_hora.copy()
df_hora_out['Fecha']    = str(fecha_hoy)
df_hora_out['Sucursal'] = df_hora_out['Nombre_Sucu'].str.upper().str.strip()
df_hora_out = df_hora_out.rename(columns={'sucursal':'ID_Sucursal','Franja_Horaria':'Hora'})
df_hora_out = df_hora_out[['ID_Sucursal','Sucursal','Fecha','Hora','Tickets','Unidades','Neto','Saldo_Cobertura','Saldo_Cliente']]
df_hora_out.to_csv(PATH_OUT_HORA, index=False, encoding='utf-8-sig')

# horas_semana_anterior.csv - detalle hora x sucursal semana pasada
df_sem_out = df_sem.copy()
df_sem_out['Fecha']    = df_sem_out['Fecha'].dt.strftime('%Y-%m-%d')
df_sem_out['Sucursal'] = df_sem_out['Nombre_Sucu'].str.upper().str.strip()
df_sem_out = df_sem_out.rename(columns={'sucursal':'ID_Sucursal','Franja_Horaria':'Hora'})
df_sem_out = df_sem_out[['ID_Sucursal','Sucursal','Fecha','Hora','Tickets','Unidades','Neto','Saldo_Cobertura','Saldo_Cliente']]
df_sem_out.to_csv(PATH_OUT_SEM, index=False, encoding='utf-8-sig')

# beneficios.csv - fila unica con datos de la red
PATH_OUT_BEN = os.path.join(BASE_DIR, "public", "data", "beneficios.csv")
pd.DataFrame([{
    'Alta_Clientes':       alta_clientes,
    'Pct_Nominados':       pct_nominados,
    'Tickets_Nominados':   total_tickets_nom,
    'Total_Tickets':       total_tickets_nom_base,
    'Promedio_Diario_Clientes': round(alta_clientes / dia_actual, 0),
    'Meta_Pct_Nominados':  35.0,
}]).to_csv(PATH_OUT_BEN, index=False, encoding='utf-8-sig')

print(f"\nLISTO - 4 archivos generados:")
print(f"  {PATH_OUT}      -> {len(df)} sucursales x {len(df.columns)} columnas")
print(f"  {PATH_OUT_HORA} -> {len(df_hora_out)} filas ({df_hora_out['Sucursal'].nunique()} sucursales x hora)")
print(f"  {PATH_OUT_SEM}  -> {len(df_sem_out)} filas ({df_sem_out['Sucursal'].nunique()} sucursales x hora)")
print(f"\n  Fecha      : {fecha_hoy} (Dia {dia_actual}/{dias_mes})")
print(f"  Franja hoy : {ultima_franja_hoy}hs")
print(f"\n  Estado acumulado:")
for estado, count in df['Estado_Acumulado'].value_counts().items():
    print(f"    {estado}: {count}")
print(f"\n  Alertas:")
for alerta, count in df['Alerta_Inactividad'].value_counts().items():
    print(f"    {alerta}: {count}")
